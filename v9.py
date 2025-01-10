from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime

# Estilos de texto
font_error = Font(color="FF0000")  # Vermelho
font_warning = Font(color="0000FF")  # Azul

# Função para processar script-logs
def process_script_logs(filepath):
    data = []
    with open(filepath, 'r') as file:
        for line in file:
            parts = line.split()
            date, time = parts[0], parts[1]
            date = datetime.strptime(date, "%b/%d/%Y").strftime("%Y-%m-%d")  # Formato YYYY-MM-DD
            script_type = parts[2]
            message = ' '.join(parts[3:])
            operator = "TIM" if "TIM" in message.upper() else "CLARO"
            font = font_error if 'error' in script_type else font_warning
            data.append((date, time, message, script_type, operator, font))
    return data

# Função para processar all-logs
def process_all_logs(filepath):
    data = []
    with open(filepath, 'r') as file:
        for line in file:
            if "netwatch" in line.lower():
                parts = line.split()
                date, time = parts[0], parts[1]
                date = datetime.strptime(date, "%b/%d/%Y").strftime("%Y-%m-%d")  # Formato YYYY-MM-DD
                status = "UP" if "up" in line.lower() else "DOWN"
                operator = "TIM" if "tim" in line.lower() else "CLARO"
                font = font_error if status == "DOWN" else font_warning
                data.append((date, time, operator, status, font))
    return data

# Processar arquivos
script_logs_data = process_script_logs('script-logs.txt.0.txt')
all_logs_data = process_all_logs('all-logs.txt.1.txt')

# Criar workbook
wb = Workbook()

# Adicionar aba para script-logs
sheet1 = wb.active
sheet1.title = "Script Logs"
headers1 = ["Data", "Hora", "Mensagem", "Fonte", "Gravidade", "Operadora", "Descrição"]
sheet1.append(headers1)

for row, (date, time, message, script_type, operator, font) in enumerate(script_logs_data, start=2):
    fonte, gravidade = script_type.split(",")  # Separar em duas colunas
    sheet1.append([date, time, message, fonte, gravidade, operator, None])
    
    # Fórmula condicional para a coluna "Descrição"
    descricao_cell = f"G{row}"
    mensagem_cell = f"C{row}"
    gravidade_cell = f"E{row}"
    sheet1[descricao_cell] = (
        f'=IF(AND(ISNUMBER(SEARCH("DOWN",{mensagem_cell})), {gravidade_cell}="error"), '
        f'"Link ou conexão perdida", '
        f'IF(AND(ISNUMBER(SEARCH("UP",{mensagem_cell})), {gravidade_cell}="warning"), '
        f'"Link ou conexão recuperado", ""))'
    )
    
    # Aplicar cor na linha inteira
    for col in range(1, 8):  # De A a G
        sheet1.cell(row=row, column=col).font = font

# Adicionar congelamento de painel e autofiltro para script-logs
sheet1.freeze_panes = "A2"
sheet1.auto_filter.ref = sheet1.dimensions

# Adicionar aba para all-logs
sheet2 = wb.create_sheet(title="All Logs")
headers2 = ["Data", "Hora", "Operadora", "Status", "Descrição"]
sheet2.append(headers2)

for row, (date, time, operator, status, font) in enumerate(all_logs_data, start=2):
    sheet2.append([date, time, operator, status, None])
    
    # Fórmula condicional para a coluna "Descrição"
    descricao_cell = f"E{row}"
    status_cell = f"D{row}"
    sheet2[descricao_cell] = (
        f'=IF({status_cell}="DOWN", "Link ou conexão perdida", '
        f'IF({status_cell}="UP", "Link ou conexão recuperado", ""))'
    )
    
    # Aplicar cor na linha inteira
    for col in range(1, 6):  # De A a E
        sheet2.cell(row=row, column=col).font = font

# Adicionar congelamento de painel e autofiltro para all-logs
sheet2.freeze_panes = "A2"
sheet2.auto_filter.ref = sheet2.dimensions

# Adicionar aba para contagem de termos
sheet3 = wb.create_sheet(title="Contagem de Termos")
headers3 = ["Termo", "Contagem"]
sheet3.append(headers3)

# Contagem de "SAIDA TIM DOWN", "SAIDA TIM UP", "SAIDA CLARO DOWN", "SAIDA CLARO UP"
terms = ["SAIDA TIM DOWN", "SAIDA TIM UP", "SAIDA CLARO DOWN", "SAIDA CLARO UP"]
count_data = []
for term in terms:
    count = sum(1 for message in [x[2] for x in script_logs_data] if term in message.upper())
    count_data.append((term, count))

# Adicionar as contagens na nova aba
for count in count_data:
    sheet3.append(count)

# Adicionar aba para contagem de combinações
sheet4 = wb.create_sheet(title="Contagem de Combinações")
headers4 = ["Combinação", "Contagem"]
sheet4.append(headers4)

# Contagem de combinações "TIM / DOWN", "TIM / UP", "CLARO / DOWN", "CLARO / UP"
combinations = ["TIM / DOWN", "TIM / UP", "CLARO / DOWN", "CLARO / UP"]
combination_data = []
for comb in combinations:
    operator, status = comb.split(" / ")
    count = sum(1 for operator_, status_ in [(x[4], x[3]) for x in all_logs_data] if operator_ == operator and status_ == status)
    combination_data.append((comb, count))

# Adicionar as contagens das combinações
for combination in combination_data:
    sheet4.append(combination)

# Ajustar largura das colunas
for sheet in [sheet1, sheet2, sheet3, sheet4]:
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# Adicionar data e hora de geração ao rodapé
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
sheet1.append([])
sheet1.append([f"Gerado em: {timestamp}"])
sheet2.append([])
sheet2.append([f"Gerado em: {timestamp}"])
sheet3.append([])
sheet3.append([f"Gerado em: {timestamp}"])
sheet4.append([])
sheet4.append([f"Gerado em: {timestamp}"])

# Salvar arquivo Excel com data/hora no nome
output_file = f"V9_saida_logs_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
wb.save(output_file)
print(f"Arquivo Excel gerado: {output_file}")
